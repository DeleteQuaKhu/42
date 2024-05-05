import os
from openpyxl import load_workbook

def rename_sheets(filename):
    base_name = os.path.splitext(os.path.basename(filename))[0]  # Extract base name without extension

    wb = load_workbook(filename)
    if 'Sheet1' in wb.sheetnames:
        ws1 = wb['Sheet1']
        ws1.title = base_name + '_R'
    if 'Sheet2' in wb.sheetnames:
        ws2 = wb['Sheet2']
        ws2.title = base_name + '_L'
    wb.save(filename)

file_path = r"C:\Users\TechnoStar\Documents\macro\save png\1234.xlsx"

rename_sheets(file_path)

