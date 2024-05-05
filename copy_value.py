import pandas as pd
from openpyxl import load_workbook
from io import StringIO

def read_csv_file(csv_file):
    # Read CSV into a DataFrame with Python's built-in CSV reader
    with open(csv_file, 'r') as file:
        lines = file.readlines()
        max_fields = max(len(line.split(',')) for line in lines)
        
        # Create a StringIO object to handle the inconsistent number of fields
        file_content = '\n'.join(line.strip() + (',' * (max_fields - len(line.split(',')))) for line in lines)
        df = pd.read_csv(StringIO(file_content), header=None)
        big_list = df.values.tolist()
    return big_list

def copy_to_excel(big_list, excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    rows_to_copy = [big_list[1], big_list[2]]

    # Copy values to cells A1:E2
    for row_index, row in enumerate(rows_to_copy, start=1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    wb.save(excel_file)


# # Specify the paths
csv_file_path = r"C:\Users\TechnoStar\Documents\macro\save png\test.csv"
excel_file_path = r"C:\Users\TechnoStar\Documents\macro\save png\1234 - Copy.xlsx"

big_list = read_csv_file(csv_file_path)
copy_to_excel(big_list, excel_file_path)

